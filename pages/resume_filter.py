"""
Resume Filter Page - Extract all content from resumes and export to Excel
"""
import streamlit as st
import pandas as pd
from datetime import datetime
import os
import sys

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agents.resume_analyzer_agent import parse_resume_with_agent
from utils.api_key_manager import get_api_key_manager

# Try to import PDF/DOCX libraries
try:
    import PyPDF2
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from docx import Document
    from docx.shared import Inches, Pt
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False
    Document = None

# BytesIO is always available (built-in)
from io import BytesIO


def extract_text_from_file(uploaded_file):
    """Extract text from PDF/DOCX/TXT files"""
    try:
        if uploaded_file.type == "application/pdf":
            # Try PyPDF2 first
            try:
                pdf_reader = PyPDF2.PdfReader(uploaded_file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() or ""
                if text.strip():
                    return text, True, None
            except:
                pass
            
            # Fallback to pdfplumber
            try:
                uploaded_file.seek(0)
                with pdfplumber.open(uploaded_file) as pdf:
                    text = ""
                    for page in pdf.pages:
                        text += page.extract_text() or ""
                if text.strip():
                    return text, True, None
            except:
                pass
            
            return "", False, "Could not extract text from PDF"
        
        elif "wordprocessingml" in uploaded_file.type or uploaded_file.name.endswith(".docx"):
            try:
                doc = Document(uploaded_file)
                text = "\n".join([para.text for para in doc.paragraphs])
                if text.strip():
                    return text, True, None
                return "", False, "DOCX file is empty"
            except Exception as e:
                return "", False, f"Error reading DOCX: {str(e)}"
        
        else:  # Text file
            try:
                text = uploaded_file.getvalue().decode("utf-8")
                if text.strip():
                    return text, True, None
                return "", False, "Text file is empty"
            except Exception as e:
                return "", False, f"Error reading text file: {str(e)}"
    
    except Exception as e:
        return "", False, f"Unexpected error: {str(e)}"


# Page config
st.set_page_config(page_title="Resume Filter", layout="wide", page_icon="🔍")

# Title
st.title("🔍 Resume Filter")
st.markdown("**Extract all information from resumes and export to Excel**")

# Check API keys
try:
    api_mgr = get_api_key_manager()
    if api_mgr.get_total_keys() > 0:
        st.success(f"✅ {api_mgr.get_total_keys()} API key(s) loaded")
    else:
        st.error("❌ No API keys loaded!")
except Exception as e:
    st.error(f"❌ Error: {str(e)}")

st.markdown("---")

# Initialize session state
if 'extracted_data' not in st.session_state:
    st.session_state.extracted_data = []

# File uploader
st.header("📤 Upload Resumes")
uploaded_files = st.file_uploader(
    "Upload multiple resumes to extract information",
    type=["pdf", "docx", "txt"],
    accept_multiple_files=True,
    help="Upload PDF, DOCX, or TXT files"
)

if uploaded_files:
    st.info(f"📊 **{len(uploaded_files)} resume(s) selected**")
    
    if st.button("🚀 Extract Information", type="primary", use_container_width=True):
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        extracted_data = []
        successful = 0
        failed = 0
        
        for idx, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Processing {idx + 1}/{len(uploaded_files)}: {uploaded_file.name}")
            
            # Extract text
            resume_text, success, error = extract_text_from_file(uploaded_file)
            
            if not success:
                st.error(f"❌ {uploaded_file.name}: {error}")
                failed += 1
                progress_bar.progress((idx + 1) / len(uploaded_files))
                continue
            
            # Parse with agent - Let agent extract ALL content dynamically
            try:
                result = parse_resume_with_agent(resume_text)
                
                if result.get("status") == "success":
                    # Agent extracts content dynamically - no predefined structure
                    # We take whatever the agent returns
                    data = {
                        "File Name": uploaded_file.name,
                        "Extracted Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    }
                    
                    # Add all fields the agent extracted (dynamic)
                    for key, value in result.items():
                        if key == "status":
                            continue
                        
                        # Format the value for display
                        if isinstance(value, list):
                            if len(value) > 0 and isinstance(value[0], dict):
                                # List of dicts (like education, experience)
                                formatted = " | ".join([str(item) for item in value])
                                data[key.replace("_", " ").title()] = formatted
                            else:
                                # Simple list (like skills)
                                data[key.replace("_", " ").title()] = ", ".join(str(v) for v in value)
                        elif isinstance(value, dict):
                            # Dict - convert to string
                            data[key.replace("_", " ").title()] = str(value)
                        else:
                            # Simple value
                            data[key.replace("_", " ").title()] = value if value else "N/A"
                    
                    extracted_data.append(data)
                    successful += 1
                    st.success(f"✅ {uploaded_file.name} - Extracted successfully")
                else:
                    st.error(f"❌ {uploaded_file.name}: {result.get('error', 'Unknown error')}")
                    failed += 1
                    
            except Exception as e:
                st.error(f"❌ {uploaded_file.name}: {str(e)}")
                failed += 1
            
            progress_bar.progress((idx + 1) / len(uploaded_files))
        
        # Save to session state
        st.session_state.extracted_data = extracted_data
        
        # Final summary
        progress_bar.progress(1.0)
        status_text.success("✅ Extraction complete!")
        
        st.success(f"""
        **Extraction Summary:**
        - ✅ Successful: {successful}
        - ❌ Failed: {failed}
        - 📊 Total: {len(uploaded_files)}
        """)

# Display extracted data
if st.session_state.extracted_data:
    st.markdown("---")
    st.header("📋 Extracted Information")
    
    # Convert to DataFrame
    df = pd.DataFrame(st.session_state.extracted_data)
    
    # Display table
    st.dataframe(df, use_container_width=True, height=400)
    
    # Download buttons
    col1, col2 = st.columns(2)
    
    with col1:
        # Excel download with formatting
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resume Data"
        
        # Add title
        ws.merge_cells('A1:' + chr(64 + len(df.columns)) + '1')
        title_cell = ws['A1']
        title_cell.value = "Resume Extraction Report"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Add data starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header row formatting
                if r_idx == 3:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    # Data rows - enable text wrapping
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Auto-size columns based on content
        from openpyxl.utils import get_column_letter
        
        for col_idx in range(1, len(df.columns) + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Check all cells in this column
            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        # Calculate length considering line breaks
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines) if lines else 0
                        max_length = max(max_length, max_line_length)
                except:
                    pass
            
            # Set column width (min 15, max 50 characters)
            adjusted_width = min(max(max_length + 2, 15), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set row heights for data rows
        for row in range(4, ws.max_row + 1):
            ws.row_dimensions[row].height = None  # Auto height
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        st.download_button(
            "📥 Download Excel",
            excel_buffer.read(),
            f"resume_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col2:
        # Clear data
        if st.button("🗑️ Clear Data", use_container_width=True):
            st.session_state.extracted_data = []
            st.rerun()
    
    # Statistics - Dynamic based on extracted data
    st.markdown("---")
    st.subheader("📊 Statistics")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Resumes", len(df))
    
    with col2:
        # Count resumes with email
        email_cols = [col for col in df.columns if 'email' in col.lower()]
        if email_cols:
            with_email = len(df[df[email_cols[0]] != "N/A"])
            st.metric("With Email", with_email)
        else:
            st.metric("Columns", len(df.columns))
    
    with col3:
        # Count resumes with phone
        phone_cols = [col for col in df.columns if 'phone' in col.lower()]
        if phone_cols:
            with_phone = len(df[df[phone_cols[0]] != "N/A"])
            st.metric("With Phone", with_phone)
        else:
            st.metric("Fields Extracted", len(df.columns) - 2)
    
    with col4:
        # Count resumes with skills
        skill_cols = [col for col in df.columns if 'skill' in col.lower()]
        if skill_cols:
            with_skills = len(df[df[skill_cols[0]] != "N/A"])
            st.metric("With Skills", with_skills)
        else:
            st.metric("Data Points", len(df) * len(df.columns))

else:
    st.info("👆 Upload resumes above to extract information")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p>🔍 Resume Filter | Powered by AI Agent + Groq Llama</p>
</div>
""", unsafe_allow_html=True)
